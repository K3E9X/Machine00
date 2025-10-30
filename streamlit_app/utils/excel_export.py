from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, Optional
import json


class ExcelExportService:
    def __init__(self, questions_data: Dict):
        self.questions_data = questions_data

    def create_questionnaire_template(self, lang: str = "fr", app_name: str = None) -> Workbook:
        """
        Create an Excel template for the questionnaire

        Args:
            lang: Language code (fr or en)
            app_name: Optional application name

        Returns:
            Workbook object
        """
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create sheets
        self._create_info_sheet(wb, lang, app_name)
        self._create_questionnaire_sheet(wb, lang)
        self._create_instructions_sheet(wb, lang)

        return wb

    def create_results_report(self, responses: Dict, score_data: Dict,
                             recommendations: list, lang: str = "fr",
                             app_info: Dict = None) -> Workbook:
        """
        Create an Excel report with results and analysis

        Args:
            responses: User responses
            score_data: Calculated score data
            recommendations: List of recommendations
            lang: Language code
            app_info: Application information

        Returns:
            Workbook object
        """
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create sheets
        self._create_summary_sheet(wb, score_data, lang, app_info)
        self._create_detailed_results_sheet(wb, responses, lang)
        self._create_recommendations_sheet(wb, recommendations, lang)

        return wb

    def _create_info_sheet(self, wb: Workbook, lang: str, app_name: Optional[str]):
        """Create information sheet"""
        ws = wb.create_sheet("Information" if lang == "en" else "Informations")

        # Title styling
        title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        header_font = Font(name="Calibri", size=11, bold=True)
        normal_font = Font(name="Calibri", size=11)

        # Title
        ws["A1"] = "Application Security Assessment Questionnaire" if lang == "en" else "Questionnaire d'Évaluation de la Sécurité Applicative"
        ws["A1"].font = title_font
        ws.merge_cells("A1:B1")

        # Application info section
        row = 3
        labels = {
            "fr": {
                "app_name": "Nom de l'application:",
                "date": "Date:",
                "owner": "Propriétaire:",
                "contact": "Contact:",
                "environment": "Environnement:",
                "description": "Description:"
            },
            "en": {
                "app_name": "Application Name:",
                "date": "Date:",
                "owner": "Owner:",
                "contact": "Contact:",
                "environment": "Environment:",
                "description": "Description:"
            }
        }

        for key, label in labels[lang].items():
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = header_font
            if key == "app_name" and app_name:
                ws[f"B{row}"] = app_name
            elif key == "date":
                ws[f"B{row}"] = datetime.now().strftime("%Y-%m-%d")
            ws[f"B{row}"].font = normal_font
            row += 1

        # Instructions
        row += 2
        instructions = {
            "fr": "Instructions:\n1. Remplissez les informations de l'application ci-dessus\n2. Accédez à l'onglet 'Questionnaire' pour répondre aux questions\n3. Sélectionnez la réponse la plus appropriée pour chaque question\n4. Retournez ce fichier complété pour analyse",
            "en": "Instructions:\n1. Fill in the application information above\n2. Go to the 'Questionnaire' tab to answer questions\n3. Select the most appropriate answer for each question\n4. Return this completed file for analysis"
        }

        ws[f"A{row}"] = instructions[lang]
        ws[f"A{row}"].font = normal_font
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"A{row}:B{row + 5}")
        ws.row_dimensions[row].height = 100

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50

    def _create_questionnaire_sheet(self, wb: Workbook, lang: str):
        """Create main questionnaire sheet"""
        ws = wb.create_sheet("Questionnaire")

        # Headers
        headers = {
            "fr": ["Catégorie", "Question", "Réponse", "Standards"],
            "en": ["Category", "Question", "Answer", "Standards"]
        }

        # Styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        category_font = Font(name="Calibri", size=11, bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col, header in enumerate(headers[lang], start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        row = 2
        for category in self.questions_data["categories"]:
            # Category header
            ws.cell(row=row, column=1, value=category["name"][lang])
            ws.cell(row=row, column=1).font = category_font
            ws.cell(row=row, column=1).fill = category_fill
            ws.merge_cells(f"A{row}:D{row}")
            row += 1

            # Questions
            for question in category["questions"]:
                ws.cell(row=row, column=1, value=category["name"][lang])
                ws.cell(row=row, column=2, value=question["text"][lang])
                ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

                # Create dropdown for answers
                options_text = "\n".join([f"{opt['label'][lang]}" for opt in question["options"]])
                ws.cell(row=row, column=3, value="")

                # Standards
                standards = ", ".join(question["standard"])
                ws.cell(row=row, column=4, value=standards)

                # Apply borders
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = border

                ws.row_dimensions[row].height = max(30, len(question["text"][lang]) / 2)
                row += 1

            row += 1  # Empty row between categories

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 25

        # Freeze panes
        ws.freeze_panes = "A2"

    def _create_instructions_sheet(self, wb: Workbook, lang: str):
        """Create instructions sheet"""
        ws = wb.create_sheet("Instructions")

        instructions_content = {
            "fr": {
                "title": "Guide d'utilisation du questionnaire",
                "sections": [
                    {
                        "header": "Objectif",
                        "content": "Ce questionnaire permet d'évaluer le niveau de sécurité de votre application selon les standards OWASP Top 10, ISO 27001 et les recommandations de l'ANSSI."
                    },
                    {
                        "header": "Comment répondre",
                        "content": "Pour chaque question:\n1. Lisez attentivement la question\n2. Sélectionnez la réponse qui correspond le mieux à votre situation\n3. En cas de doute, choisissez la réponse la plus conservatrice\n4. N'hésitez pas à consulter votre équipe technique"
                    },
                    {
                        "header": "Catégories couvertes",
                        "content": "1. Gestion des Identités et des Accès (IAM)\n2. Architecture Réseau et Segmentation\n3. Flux et Interconnexions\n4. Hébergement et Infrastructure\n5. Sécurité des Données\n6. Sécurité Applicative\n7. Journalisation et Surveillance"
                    },
                    {
                        "header": "Résultats",
                        "content": "À l'issue de l'évaluation, vous recevrez:\n- Un score global de sécurité\n- Une analyse par catégorie\n- Des recommandations d'amélioration\n- Une préconisation d'audit (complet, ciblé ou léger)"
                    }
                ]
            },
            "en": {
                "title": "Questionnaire User Guide",
                "sections": [
                    {
                        "header": "Objective",
                        "content": "This questionnaire assesses your application's security level according to OWASP Top 10, ISO 27001 standards and ANSSI recommendations."
                    },
                    {
                        "header": "How to Answer",
                        "content": "For each question:\n1. Read the question carefully\n2. Select the answer that best matches your situation\n3. When in doubt, choose the most conservative answer\n4. Feel free to consult your technical team"
                    },
                    {
                        "header": "Covered Categories",
                        "content": "1. Identity and Access Management (IAM)\n2. Network Architecture and Segmentation\n3. Flows and Interconnections\n4. Hosting and Infrastructure\n5. Data Security\n6. Application Security\n7. Logging and Monitoring"
                    },
                    {
                        "header": "Results",
                        "content": "Upon completion, you will receive:\n- An overall security score\n- Analysis by category\n- Improvement recommendations\n- Audit recommendation (full, targeted, or light)"
                    }
                ]
            }
        }

        # Title
        ws["A1"] = instructions_content[lang]["title"]
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        ws.merge_cells("A1:B1")

        row = 3
        for section in instructions_content[lang]["sections"]:
            ws[f"A{row}"] = section["header"]
            ws[f"A{row}"].font = Font(name="Calibri", size=12, bold=True)
            row += 1

            ws[f"A{row}"] = section["content"]
            ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(f"A{row}:B{row + 3}")
            ws.row_dimensions[row].height = 80
            row += 5

        ws.column_dimensions["A"].width = 80

    def _create_summary_sheet(self, wb: Workbook, score_data: Dict, lang: str, app_info: Dict):
        """Create summary results sheet"""
        ws = wb.create_sheet("Summary" if lang == "en" else "Résumé", 0)

        # Title
        title = "Security Assessment Summary" if lang == "en" else "Résumé de l'Évaluation de Sécurité"
        ws["A1"] = title
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        ws.merge_cells("A1:D1")

        # Overall score
        row = 3
        ws[f"A{row}"] = "Overall Score" if lang == "en" else "Score Global"
        ws[f"A{row}"].font = Font(bold=True, size=12)

        ws[f"B{row}"] = f"{score_data['percentage']}%"
        ws[f"B{row}"].font = Font(bold=True, size=20, color=score_data['risk_level']['color'].replace('#', ''))

        # Risk level
        row += 1
        ws[f"A{row}"] = "Risk Level" if lang == "en" else "Niveau de Risque"
        ws[f"A{row}"].font = Font(bold=True)

        ws[f"B{row}"] = score_data['risk_level'][lang]
        ws[f"B{row}"].font = Font(bold=True, color=score_data['risk_level']['color'].replace('#', ''))

        # Audit recommendation
        row += 1
        ws[f"A{row}"] = "Recommendation" if lang == "en" else "Recommandation"
        ws[f"A{row}"].font = Font(bold=True)

        ws[f"B{row}"] = score_data['audit_recommendation'][lang]
        ws[f"B{row}"].font = Font(bold=True)

        # Category breakdown
        row += 3
        ws[f"A{row}"] = "Category Breakdown" if lang == "en" else "Détail par Catégorie"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        # Headers for category table
        headers = ["Category" if lang == "en" else "Catégorie",
                   "Score",
                   "Percentage" if lang == "en" else "Pourcentage"]

        for col, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="D9E1F2",
                                                              end_color="D9E1F2",
                                                              fill_type="solid")

        row += 1

        # Category data
        for cat_id, cat_score in score_data['category_scores'].items():
            # Find category name
            cat_name = ""
            for cat in self.questions_data["categories"]:
                if cat["id"] == cat_id:
                    cat_name = cat["name"][lang]
                    break

            ws.cell(row=row, column=1, value=cat_name)
            ws.cell(row=row, column=2, value=f"{cat_score['score']}/{cat_score['max_score']}")
            ws.cell(row=row, column=3, value=f"{cat_score['percentage']}%")
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _create_detailed_results_sheet(self, wb: Workbook, responses: Dict, lang: str):
        """Create detailed results sheet"""
        ws = wb.create_sheet("Detailed Results" if lang == "en" else "Résultats Détaillés")

        # Headers
        headers = {
            "fr": ["Catégorie", "Question", "Réponse Sélectionnée", "Score"],
            "en": ["Category", "Question", "Selected Answer", "Score"]
        }

        for col, header in enumerate(headers[lang], start=1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).fill = PatternFill(start_color="1F4E78",
                                                           end_color="1F4E78",
                                                           fill_type="solid")
            ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")

        row = 2
        for category in self.questions_data["categories"]:
            for question in category["questions"]:
                if question["id"] in responses:
                    user_value = responses[question["id"]]

                    # Find selected option
                    selected_label = ""
                    for opt in question["options"]:
                        if opt["value"] == user_value:
                            selected_label = opt["label"][lang]
                            break

                    ws.cell(row=row, column=1, value=category["name"][lang])
                    ws.cell(row=row, column=2, value=question["text"][lang])
                    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
                    ws.cell(row=row, column=3, value=selected_label)
                    ws.cell(row=row, column=4, value=user_value)

                    row += 1

        # Adjust columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 10

    def _create_recommendations_sheet(self, wb: Workbook, recommendations: list, lang: str):
        """Create recommendations sheet"""
        ws = wb.create_sheet("Recommendations" if lang == "en" else "Recommandations")

        title = "Improvement Recommendations" if lang == "en" else "Recommandations d'Amélioration"
        ws["A1"] = title
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
        ws.merge_cells("A1:D1")

        row = 3
        headers = {
            "fr": ["Sévérité", "Catégorie", "Question", "Standards"],
            "en": ["Severity", "Category", "Question", "Standards"]
        }

        for col, header in enumerate(headers[lang], start=1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).font = Font(bold=True)

        row += 1

        for rec in recommendations:
            severity_color = "ef4444" if rec["severity"] == "high" else "f59e0b"

            ws.cell(row=row, column=1, value=rec["severity"].upper())
            ws.cell(row=row, column=1).font = Font(color=severity_color, bold=True)

            ws.cell(row=row, column=2, value=rec["category"])
            ws.cell(row=row, column=3, value=rec["question"])
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=4, value=", ".join(rec["standard"]))

            row += 1

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 25
